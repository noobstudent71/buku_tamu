import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side
from .forms import TamuForm
from .models import BukuTamu
from collections import Counter 
import json
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.db.models import Case, When, Value, IntegerField
from django.core.paginator import Paginator 
from django.db.models import Q
from django.db.models import Count
from django.db.models.functions import ExtractMonth
from django.utils.timezone import localdate, make_aware
from .models import PIC, Instansi, BukuTamu
from django.contrib.auth import logout
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse
from django.contrib.auth.models import User, Group
from .models import Instansi, PIC, BukuTamu, LogAktivitas
from django.db import IntegrityError
from django.contrib.auth.signals import user_logged_in, user_logged_out
from django.dispatch import receiver
from django.contrib.auth.models import User, Group


def search_perusahaan(request):
    # Tangkap huruf yang sedang diketik user (misal: "Otsu")
    query = request.GET.get('q', '')
    
    if query:
        # Cari di database: Perusahaan yang mengandung kata (icontains) dari query
        # values_list flat=True: Ambil namanya saja, distinct(): Jangan ada nama dobel
        # [:10] = Batasi maksimal 10 saran agar server tidak berat
        hasil = Instansi.objects.filter(nama_standar__icontains=query).values_list('nama_standar', flat=True).distinct()[:10]
        data = list(hasil)
    else:
        data = []
        
    # Kembalikan jawaban dalam bentuk JSON (Bahasa universal antar sistem)
    return JsonResponse(data, safe=False)


def login_view(request):
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        
        user = authenticate(request, username=u, password=p)
        
        if user is not None:
            login(request, user)
            
            # --- 1. CEK SURAT PENGANTAR (?next=) ---
            # Jika URL memiliki akhiran ?next=/master-data/, tangkap tujuannya
            next_url = request.GET.get('next') 
            if next_url:
                return redirect(next_url) # Langsung luncurkan ke halaman yang dia tuju!
            
            # --- 2. JIKA LOGIN BIASA (TANPA ?next=), PAKAI JALUR DEFAULT ---
            if user.groups.filter(name='SATPAM').exists():
                return redirect('daftar_tamu') 
            elif user.groups.filter(name='KANTOR').exists() or user.is_superuser:
                return redirect('dashboard_analytics') 
            else:
                return redirect('pendaftaran_tamu') 
        else:
            messages.error(request, "Username atau Password salah!")
            
    return render(request, 'tamu/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')


def form_tamu(request):
    if request.method == "POST":
        # Ambil data dari form
        nama_tamu = request.POST.get('nama')
        nama_instansi = request.POST.get('instansi')
        nama_pic = request.POST.get('pic_tuju') # Ini akan menerima string nama

        # Cari objek PIC berdasarkan nama yang diketik
        pic_obj = PIC.objects.filter(nama_lengkap=nama_pic, is_active=True).first()

        # Simpan ke database
        BukuTamu.objects.create(
            nama=nama_tamu,
            instansi=nama_instansi,
            pic_tuju=pic_obj, # Django akan otomatis mengambil ID-nya
            # ... tambahkan kolom lain sesuai modelmu ...
        )
        return redirect('halaman_sukses')

    # Ambil semua data untuk rekomendasi (datalist)
    daftar_pic = PIC.objects.filter(is_active=True)
    daftar_instansi = Instansi.objects.all()

    context = {
        'daftar_pic': daftar_pic,
        'daftar_instansi': daftar_instansi,
    }
    return render(request, 'tamu/form_tamu.html', context)


@login_required(login_url='login')
def arsip_tamu(request):
    # 1. Ambil SEMUA data (urutkan dari yang terbaru)
    tamu_list = BukuTamu.objects.all().order_by('-created_at')

    # 2. Tangkap data dari Filter UI (Kita sisakan 2 saja: pencarian universal & tanggal)
    cari_nama = request.GET.get('nama', '')  # Ini sekarang bertugas sebagai penangkap Omni-Search
    cari_tanggal = request.GET.get('tanggal', '')

    # 3. Eksekusi Filter
    if cari_nama:
        from django.db.models import Q
        # Logika Sapu Jagat: Cari di semua kolom sekaligus
        tamu_list = tamu_list.filter(
            Q(nama__icontains=cari_nama) | 
            Q(no_hp__icontains=cari_nama) |
            Q(instansi__icontains=cari_nama) |
            Q(pic_tuju__nama_lengkap__icontains=cari_nama) |
            Q(pic_tuju__departemen__icontains=cari_nama)
        )
        
    if cari_tanggal:
        tamu_list = tamu_list.filter(waktu_masuk__date=cari_tanggal)

    # 4. Fitur Pagination (20 data per halaman)
    paginator = Paginator(tamu_list, 20) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'cari_nama': cari_nama, # Tetap pakai variabel ini agar HTML tidak error
        'cari_tanggal': cari_tanggal,
    }
    return render(request, 'tamu/arsip_tamu.html', context)

# --- VIEW 1: Form Pendaftaran (Di sinilah Langkah 2 dipasang!) ---
def pendaftaran_tamu(request):
    if request.method == 'POST':
        # 1. Bikin 'copy' dari data form agar bisa kita edit di tengah jalan
        data_post = request.POST.copy()
        
        # 2. Tangkap ketikan tamu (Contoh: "Budi - HRD")
        nama_pic_input = request.POST.get('pic_tuju', '')
        pic_obj = None
        
        # 3. --- LOGIKA LANGKAH 2: POTONG NAMA DAN DEPARTEMEN ---
        if nama_pic_input:
            if " - " in nama_pic_input:
                nama_asli, dept_asli = nama_pic_input.split(" - ", 1)
                pic_obj = PIC.objects.filter(nama_lengkap=nama_asli.strip(), departemen=dept_asli.strip(), is_active=True).first()
            else:
                pic_obj = PIC.objects.filter(nama_lengkap=nama_pic_input.strip(), is_active=True).first()
        
        # 4. Jika PIC ketemu, ganti teks "Budi - HRD" menjadi ID-nya Pak Budi
        if pic_obj:
            data_post['pic_tuju'] = pic_obj.id

        # 5. Lanjutkan proses save form seperti biasa dengan data yang sudah di-"hack"
        form = TamuForm(data_post, request.FILES)
        
        if form.is_valid():
            tamu = form.save(commit=False)
            
            # --- LOGIKA STATUS ---
            if tamu.sudah_janji:
                tamu.status = 'MASUK'
                messages.success(request, f"Selamat Datang, {tamu.nama}! Silakan langsung masuk.")
            else:
                tamu.status = 'MENUNGGU'
                messages.warning(request, f"Terima Kasih, {tamu.nama}. Mohon tunggu konfirmasi Security.")
            
            tamu.save()
            return redirect('pendaftaran_tamu')
        else:
            messages.error(request, "Gagal menyimpan data! Pastikan Nama PIC dipilih dari daftar.")
            
    else:
        form = TamuForm()

    # ---> AMBIL DATA DARI DATABASE UNTUK REKOMENDASI (DATALIST) <---
    daftar_pic = PIC.objects.filter(is_active=True)
    daftar_instansi = Instansi.objects.all()

    context = {
        'form': form,
        'daftar_pic': daftar_pic,
        'daftar_instansi': daftar_instansi,
    }

    return render(request, 'tamu/form_tamu.html', context)

# --- VIEW 2: Dashboard Monitor (Kode Anda) ---
@login_required(login_url='login')
def daftar_tamu(request):
  
    base_query = BukuTamu.objects.annotate(
        urutan_prioritas=Case(
            When(status='MENUNGGU', then=Value(1)), # Prioritas 1: Menunggu
            When(status='MASUK', then=Value(2)),    # Prioritas 2: Sedang Bertamu
            default=Value(3),                       # Prioritas 3: Selesai
            output_field=IntegerField(),
        )
    ).order_by('urutan_prioritas', '-waktu_masuk')

    # 2. Logika Filter Tombol (Jika diklik)
    status_filter = request.GET.get('status')
    
    if status_filter == 'masuk':
        tamu_list = base_query.filter(status='MASUK')
    elif status_filter == 'keluar':
        tamu_list = base_query.filter(status='KELUAR')
    else:
        # Jika tidak ada filter, tampilkan campuran
        tamu_list = base_query

    # 3. [PENTING] BATASI HANYA 50 DATA TERAKHIR
    #    Ini yang membuat dashboard tetap ringan dan rapi.
    #    Sisanya? Silakan cari di menu Arsip.
    tamu_list = tamu_list[:50]

    # 4. Hitung Statistik (Global Real-time)
    #    Angka statistik tetap menghitung TOTAL SELURUHNYA di gedung,
    #    tidak peduli apakah dia masuk di list 50 atau tidak.
    stat_masuk = BukuTamu.objects.filter(status='MASUK').count()
    stat_pulang = BukuTamu.objects.filter(status='KELUAR').count()
    total_arsip = BukuTamu.objects.count()

    context = {
        'tamu_list': tamu_list,
        'sedang_bertamu': stat_masuk,
        'sudah_pulang': stat_pulang,
        'total_arsip': total_arsip,
    }
    
    return render(request, 'tamu/daftar_tamu.html', context)


# --- VIEW 3: Ubah Status (Kode Anda) ---
@login_required(login_url='login')
def ubah_status(request, id, status_baru):
    tamu = get_object_or_404(BukuTamu, id=id)
    tamu.status = status_baru
    
    # Jika tamu keluar ATAU ditolak masuk, catat jam keluarnya saat ini juga
    if status_baru == 'KELUAR' or status_baru == 'DITOLAK':
        tamu.waktu_keluar = timezone.now()
        
    tamu.save()
    
    # Notifikasi yang menyesuaikan
    if status_baru == 'DITOLAK':
        messages.error(request, f'Kedatangan {tamu.nama} telah ditolak.')
    else:
        messages.success(request, f'Status {tamu.nama} diperbarui menjadi {status_baru}')
        
    return redirect('daftar_tamu')


# --- VIEW 4: DOWNLOAD EXCEL (MULTI-SHEET) ---
def download_excel(request):
    # 1. Ambil input tanggal
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        messages.warning(request, "⚠️ Silakan pilih rentang tanggal terlebih dahulu!")
        return redirect('arsip_tamu')

    try:
        start_dt = make_aware(datetime.strptime(f"{start_date} 00:00:00", "%Y-%m-%d %H:%M:%S"))
        end_dt = make_aware(datetime.strptime(f"{end_date} 23:59:59", "%Y-%m-%d %H:%M:%S"))
    except ValueError:
        messages.error(request, "Format tanggal tidak valid.")
        return redirect('arsip_tamu')


    # 2. Siapkan Workbook Excel
    wb = openpyxl.Workbook()
    
    # Hapus sheet default "Sheet" agar bersih, nanti kita buat baru
    default_ws = wb.active
    wb.remove(default_ws)

    # 3. Definisi Kategori untuk 4 Sheet
    # Format: ('value_di_db', 'Judul Sheet')
    kategori_list = [
        ('rekan_bisnis', 'LOKAL (BISNIS)'),
        ('pemerintahan', 'PEMERINTAHAN'),
        ('akademisi', 'AKADEMISI'),
        ('overseas', 'LUAR NEGERI'),
    ]

    # Style Border
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 4. LOOPING MEMBUAT 4 SHEET
    for db_value, sheet_title in kategori_list:
        
        # A. Buat Sheet Baru
        ws = wb.create_sheet(title=sheet_title)

        # B. Filter Data Spesifik per Kategori
        data_tamu = BukuTamu.objects.filter(
            created_at__date__range=[start_date, end_date],
            kategori_tamu=db_value  # <--- Filter Kunci
        ).order_by('created_at')

        # C. Header Judul Besar
        judul = f"LAPORAN TAMU KATEGORI {sheet_title} ({start_date} S/D {end_date})"
        ws.merge_cells('A1:S1') 
        ws['A1'] = judul
        ws['A1'].font = Font(size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # D. Header Kolom
        headers = [
            "NO", "NO. TIKET", "KATEGORI", "TANGGAL", "JAM MASUK", "JAM KELUAR",
            "NAMA LENGKAP", "NO. TELFON", "INSTANSI", "NOPOL", "BERTEMU SIAPA", 
            "JML ORG", "KEPERLUAN", "SUDAH JANJI?", "BARANG BAWAAN", "JML BARANG",
        ]

        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header_title
            cell.font = Font(bold=True, color="FFFFFF")
            # Warna Biru Otsuka
            cell.fill = openpyxl.styles.PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid") 
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # E. Isi Data (Jika kosong, loop ini tidak jalan, jadi tabel kosong)
        row_num = 3
        
        for index, tamu in enumerate(data_tamu, 1):
            # Ambil Label Kategori
            kategori_label = tamu.get_kategori_tamu_display()
            
            # Format Waktu
            waktu_masuk_wib = timezone.localtime(tamu.waktu_masuk) 
            tanggal_saja = waktu_masuk_wib.strftime("%d-%m-%Y")
            jam_masuk = waktu_masuk_wib.strftime("%H:%M")
            jam_keluar = timezone.localtime(tamu.waktu_keluar).strftime("%H:%M") if tamu.waktu_keluar else "-"
            
            status_janji = "YA" if tamu.sudah_janji else "TIDAK"
            nama_ktp = os.path.basename(tamu.foto_ktp.name) if tamu.foto_ktp else "-"
            nama_selfie = os.path.basename(tamu.foto_wajah.name) if tamu.foto_wajah else "-"

            nama_pic_excel = tamu.pic_tuju.nama_lengkap if tamu.pic_tuju else "-"

            row_data = [
                index, tamu.nomor_tiket, kategori_label, tanggal_saja, jam_masuk, jam_keluar,
                tamu.nama, tamu.no_hp, tamu.instansi, tamu.no_polisi, 
                nama_pic_excel, tamu.jumlah_tamu, tamu.keperluan, # <--- SUDAH AMAN
                status_janji, tamu.bawa_barang, tamu.jumlah_barang,
            ]

            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            row_num += 1

        # F. Atur Lebar Kolom (Biar Rapi)
        column_widths = [5, 18, 20, 15, 20, 20, 25, 15, 25, 15, 20, 8, 25, 20, 25, 20, 25, 25, 35]
        for i, width in enumerate(column_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

    # 5. Output File
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Laporan_Tamu_Otsuka_{start_date}.xlsx'
    wb.save(response)
    # Catat ke Audit Trail
   # Catat ke Audit Trail
    LogAktivitas.objects.create(user=request.user, aksi="EXPORT", target="Laporan Analytics (Excel)")
    
    return response

def bersihkan_nama_instansi(teks_mentah):
    if not teks_mentah:
        return "LAIN-LAIN"
    
    # 1. Kecilkan huruf inputan tamu untuk dicocokkan
    input_bersih = teks_mentah.lower().strip()
    
    # 2. Ambil SEMUA data instansi & kata kuncinya dari Database Admin
    semua_instansi = Instansi.objects.all()
    
    for item in semua_instansi:
        # Pecah kata kunci dari admin (contoh: "ub, brawijaya") menjadi list
        list_keyword = [k.strip().lower() for k in item.kata_kunci.split(',')]
        
        # Cek apakah input tamu sama persis atau mengandung salah satu kata kunci
        for keyword in list_keyword:
            if keyword == input_bersih or keyword in input_bersih:
                return item.nama_standar # Kembalikan nama paten dari database
    
    # 3. Jika Instansi BARU (Belum ada di Admin), kembalikan nama aslinya tapi dibesarkan hurufnya
    return teks_mentah.upper().strip()
  

# --- VIEW 5: DASHBOARD ANALITIK (GRAFIK) ---
@login_required(login_url='login') 
def dashboard_analytics(request):
    # --- GEMBOK KEAMANAN KANTOR ---
    if not (request.user.is_superuser or request.user.groups.filter(name='KANTOR').exists()):
        messages.error(request, "AKSES DITOLAK: Halaman Analytics khusus untuk Staff Kantor.")
        return redirect('daftar_tamu')
    # ------------------------------
    from django.utils import timezone 
    
    # 1. Ambil tahun dari request (Default: tahun saat ini)
    tahun_sekarang = timezone.now().year
    tahun_dipilih = int(request.GET.get('tahun', tahun_sekarang))
    
    # 2. Ambil daftar tahun unik yang ada di database untuk menu Dropdown
    daftar_tahun = BukuTamu.objects.dates('waktu_masuk', 'year', order='DESC')
    list_tahun = [dt.year for dt in daftar_tahun]
    
    # Pastikan tahun sekarang masuk list jika database masih kosong
    if tahun_sekarang not in list_tahun:
        list_tahun.append(tahun_sekarang)

    # ==========================================
    # 3. DATA LINE CHART (TREN BULANAN)
    # ==========================================
    tamu_per_bulan = BukuTamu.objects.filter(waktu_masuk__year=tahun_dipilih)\
        .annotate(bulan=ExtractMonth('waktu_masuk'))\
        .values('bulan')\
        .annotate(total=Count('id'))\
        .order_by('bulan')
    
    data_bulan = [0] * 12 
    for item in tamu_per_bulan:
        data_bulan[item['bulan'] - 1] = item['total']

    # ==========================================
    # 4. DATA PIE CHART (KOMPOSISI KATEGORI)
    # ==========================================
    # Tambahkan filter tahun_dipilih agar Pie Chart ikut dinamis
    tamu_per_kategori = BukuTamu.objects.filter(waktu_masuk__year=tahun_dipilih)\
        .values('kategori_tamu')\
        .annotate(total=Count('id'))\
        .order_by('-total')
        
    labels_kategori = []
    data_kategori = []
    
    for item in tamu_per_kategori:
        # Ambil nama cantik dari KATEGORI_CHOICES di models.py
        label_cantik = dict(BukuTamu.KATEGORI_CHOICES).get(item['kategori_tamu'], item['kategori_tamu'])
        labels_kategori.append(label_cantik)
        data_kategori.append(item['total'])

    # ==========================================
    # 5. DATA BAR CHART (TOP 5 INSTANSI TERPOPULER)
    # ==========================================
    # Tambahkan filter tahun_dipilih agar Bar Chart ikut dinamis
    semua_instansi = BukuTamu.objects.filter(waktu_masuk__year=tahun_dipilih)\
        .values_list('instansi', flat=True)
    
    # Bersihkan nama instansi pakai fungsi helper yang sudah Mas buat sebelumnya
    instansi_bersih = [bersihkan_nama_instansi(nama) for nama in semua_instansi]
    
    # Hitung jumlahnya pakai Counter
    hitungan_instansi = Counter(instansi_bersih)
    
    # Ambil 5 Terbanyak
    top_5 = hitungan_instansi.most_common(5)
    
    labels_instansi = [item[0] for item in top_5]
    data_instansi = [item[1] for item in top_5]

    # ==========================================
    # 6. KIRIM KE TEMPLATE HTML
    # ==========================================
    context = {
        'data_bulan': json.dumps(data_bulan),
        'labels_kategori': json.dumps(labels_kategori),
        'data_kategori': json.dumps(data_kategori),
        'labels_instansi': json.dumps(labels_instansi),
        'data_instansi': json.dumps(data_instansi),
        
        # Variabel untuk Dropdown Filter
        'tahun_aktif': tahun_dipilih, 
        'list_tahun': list_tahun,      
    }

    return render(request, 'tamu/dashboard_analytics.html', context)

@login_required(login_url='login')
def master_data(request):
    # --- GEMBOK KEAMANAN KANTOR ---
    if not (request.user.is_superuser or request.user.groups.filter(name='KANTOR').exists()):
        messages.error(request, "AKSES DITOLAK: Halaman Master Data khusus untuk Staff Kantor.")
        return redirect('daftar_tamu')
    # ------------------------------

    if request.method == 'POST':
        tipe_data = request.POST.get('tipe_data')
        
        if tipe_data == 'akun':
            username = request.POST.get('username')
            password = request.POST.get('password')
            role = request.POST.get('role')
            if username and password:
                if User.objects.filter(username=username).exists():
                    messages.error(request, f"Username '{username}' sudah terdaftar!")
                else:
                    user_baru = User.objects.create_user(username=username, password=password)
                    group, created = Group.objects.get_or_create(name=role)
                    user_baru.groups.add(group)
                    # --- REKAM LOG ---
                    LogAktivitas.objects.create(user=request.user, aksi="TAMBAH", target=f"Akun Login: {username} ({role})")
                    messages.success(request, f"Akun {role} '{username}' berhasil dibuat!")

        elif tipe_data == 'instansi':
            nama_baru = request.POST.get('nama_standar')
            kata_kunci = request.POST.get('kata_kunci', '')
            if nama_baru:
                Instansi.objects.create(nama_standar=nama_baru, kata_kunci=kata_kunci)
                # --- REKAM LOG ---
                LogAktivitas.objects.create(user=request.user, aksi="TAMBAH", target=f"Instansi: {nama_baru}")
                messages.success(request, f"Instansi '{nama_baru}' berhasil ditambahkan!")
                
        elif tipe_data == 'pic':
            nama_baru = request.POST.get('nama_lengkap')
            dept_baru = request.POST.get('departemen', 'Umum')
            if nama_baru:
                PIC.objects.create(nama_lengkap=nama_baru, departemen=dept_baru, is_active=True)
                # --- REKAM LOG ---
                LogAktivitas.objects.create(user=request.user, aksi="TAMBAH", target=f"PIC: {nama_baru} ({dept_baru})")
                messages.success(request, f"PIC '{nama_baru}' ({dept_baru}) berhasil ditambahkan!")

        # === TAMBAHKAN BLOK KODE INI UNTUK FITUR EDIT ===
        elif tipe_data == 'edit_pic':
            pic_id = request.POST.get('pic_id')
            nama_baru = request.POST.get('nama_lengkap')
            dept_baru = request.POST.get('departemen', 'UMUM')
            
            if pic_id and nama_baru:
                pic_obj = get_object_or_404(PIC, id=pic_id)
                nama_lama = pic_obj.nama_lengkap
                
                # Update datanya
                pic_obj.nama_lengkap = nama_baru
                pic_obj.departemen = dept_baru
                pic_obj.save()
                
            # --- KODE BARU UNTUK EDIT INSTANSI ---
        elif tipe_data == 'edit_instansi':
            instansi_id = request.POST.get('instansi_id')
            nama_baru = request.POST.get('nama_standar')
            kunci_baru = request.POST.get('kata_kunci', '')
            
            if instansi_id and nama_baru:
                ins_obj = get_object_or_404(Instansi, id=instansi_id) # Sesuaikan nama Model Instansi-nya
                nama_lama = ins_obj.nama_standar
                
                # Update data
                ins_obj.nama_standar = nama_baru
                ins_obj.kata_kunci = kunci_baru
                ins_obj.save()
                
                # Catat ke Audit Trail
                LogAktivitas.objects.create(user=request.user, aksi="UBAH", target=f"Instansi: {nama_lama} -> {nama_baru}")
                messages.success(request, f"Instansi '{nama_baru}' berhasil diperbarui!")

        # --- KODE BARU UNTUK EDIT AKUN LOGIN ---
        elif tipe_data == 'edit_akun':
            akun_id = request.POST.get('akun_id')
            username_baru = request.POST.get('username')
            password_baru = request.POST.get('password')
            role_baru = request.POST.get('role')
            
            if akun_id and username_baru:
                user_obj = get_object_or_404(User, id=akun_id)
                username_lama = user_obj.username
                
                # Update username
                user_obj.username = username_baru
                
                # Update password HANYA JIKA diisi
                if password_baru:
                    user_obj.set_password(password_baru)
                
                user_obj.save()
                
                # Update Role (Hak Akses)
                if role_baru:
                    user_obj.groups.clear() # Hapus role lama
                    group, created = Group.objects.get_or_create(name=role_baru)
                    user_obj.groups.add(group)
                
                # Catat ke Audit Trail
                LogAktivitas.objects.create(user=request.user, aksi="UBAH", target=f"Akun: {username_lama} -> {username_baru} ({role_baru})")
                messages.success(request, f"Data Akun '{username_baru}' berhasil diperbarui!")   

        return redirect('master_data')

    context = {
        'daftar_instansi': Instansi.objects.all().order_by('nama_standar'),
        'daftar_pic': PIC.objects.all().order_by('nama_lengkap'),
        'daftar_user': User.objects.all().order_by('-date_joined'),
    }
    return render(request, 'tamu/master_data.html', context)

# --- VIEW 7: HAPUS DATA MASTER ---
@login_required(login_url='login')
def hapus_data(request, tipe, id):
    # --- GEMBOK KEAMANAN KANTOR ---
    if not (request.user.is_superuser or request.user.groups.filter(name='KANTOR').exists()):
        messages.error(request, "AKSES DITOLAK: Halaman Master Data khusus untuk Staff Kantor.")
        return redirect('daftar_tamu')
    # ------------------------------
    if tipe == 'instansi':
        obj = get_object_or_404(Instansi, id=id)
        nama = obj.nama_standar
        obj.delete()
        LogAktivitas.objects.create(user=request.user, aksi="HAPUS", target=f"Instansi: {nama}")
        messages.success(request, f"Data '{nama}' berhasil dihapus.")
        
    elif tipe == 'pic':
        obj = get_object_or_404(PIC, id=id)
        nama = obj.nama_lengkap
        obj.delete()
        LogAktivitas.objects.create(user=request.user, aksi="HAPUS", target=f"PIC: {nama}")
        messages.success(request, f"Data '{nama}' berhasil dihapus.")
        
    elif tipe == 'akun':
        obj = get_object_or_404(User, id=id)
        nama = obj.username
        if obj.is_superuser:
            messages.error(request, "Akun Super Admin tidak boleh dihapus!")
            return redirect('master_data')
            
        # PENGAMANAN ERROR DATABASE (TRY-EXCEPT)
        try:
            obj.delete()
            LogAktivitas.objects.create(user=request.user, aksi="HAPUS", target=f"Akun Login: {nama}")
            messages.success(request, f"Akun '{nama}' berhasil dihapus permanen.")
        except IntegrityError:
            # Jika database menolak karena akun sudah punya sejarah
            messages.error(request, f"GAGAL: Akun '{nama}' tidak bisa dihapus karena memiliki riwayat aktivitas. Silakan gunakan tombol NONAKTIFKAN!")
            
    return redirect('master_data')

# --- VIEW 8: UBAH STATUS AKTIF/NONAKTIF ---
@login_required(login_url='login')
def toggle_status(request, tipe, id):
    if tipe == 'pic':
        obj = get_object_or_404(PIC, id=id)
        obj.is_active = not obj.is_active
        obj.save()
        status_str = "Diaktifkan" if obj.is_active else "Dinonaktifkan"
        LogAktivitas.objects.create(user=request.user, aksi="UBAH STATUS", target=f"PIC: {obj.nama_lengkap} -> {status_str}")
        messages.success(request, f"Status PIC '{obj.nama_lengkap}' berhasil {status_str}.")
        
    elif tipe == 'akun':
        obj = get_object_or_404(User, id=id)
        if obj.is_superuser:
            messages.error(request, "Status Super Admin tidak boleh diubah!")
            return redirect('master_data')
        obj.is_active = not obj.is_active
        obj.save()
        status_str = "Diaktifkan" if obj.is_active else "Dinonaktifkan"
        LogAktivitas.objects.create(user=request.user, aksi="UBAH STATUS", target=f"Akun Login: {obj.username} -> {status_str}")
        messages.success(request, f"Akun '{obj.username}' berhasil {status_str}.")
        
    return redirect('master_data')

# --- VIEW 9: HALAMAN AUDIT TRAIL ---
from django.core.paginator import Paginator
@login_required(login_url='login')
def audit_trail(request):
    if not (request.user.is_superuser or request.user.groups.filter(name='KANTOR').exists()):
        messages.error(request, "AKSES DITOLAK.")
        return redirect('daftar_tamu')

    logs_list = LogAktivitas.objects.all().order_by('-waktu')

    # 1. Tangkap data dari form pencarian
    cari_user = request.GET.get('user', '')
    cari_aksi = request.GET.get('aksi', 'Semua')
    cari_tanggal = request.GET.get('tanggal', '')

    # 2. Eksekusi Filter
    if cari_user:
        logs_list = logs_list.filter(user__username__icontains=cari_user)
    if cari_aksi != 'Semua':
        logs_list = logs_list.filter(aksi=cari_aksi)
    if cari_tanggal:
        logs_list = logs_list.filter(waktu__date=cari_tanggal)

    # 3. Paginasi
    paginator = Paginator(logs_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj, 
        'cari_user': cari_user, 
        'cari_aksi': cari_aksi,
        'cari_tanggal': cari_tanggal, 
    }
    return render(request, 'tamu/audit_trail.html', context)    

# --- SENSOR OTOMATIS LOGIN & LOGOUT ---
@receiver(user_logged_in)
def catat_login(sender, request, user, **kwargs):
    LogAktivitas.objects.create(user=user, aksi="LOGIN", target="-")

@receiver(user_logged_out)
def catat_logout(sender, request, user, **kwargs):
    LogAktivitas.objects.create(user=user, aksi="LOGOUT", target="-")
    